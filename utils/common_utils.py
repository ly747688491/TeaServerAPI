"""
@Project        ：tea_server_api
@File           ：common_utils.py
@IDE            ：PyCharm
@Author         ：李延
@Date           ：2024/5/7 下午6:00
@Description    ：
"""

import io
import os
import re
import uuid
from datetime import datetime
from typing import AnyStr, Dict, List, Optional, TypeVar
from urllib.parse import urlparse

from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel

from config import setting


def bytes2human(n, format_str="%(value).1f%(symbol)s"):
    """Used by various scripts. See:
    https://goo.gl/zeJZl

    >>> bytes2human(10000)
    '9.8K'
    >>> bytes2human(100001221)
    '95.4M'
    """
    symbols = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    prefix = {s: 1 << (i + 1) * 10 for i, s in enumerate(symbols[1:])}
    for symbol in reversed(symbols[1:]):
        if n >= prefix[symbol]:
            value = float(n) / prefix[symbol]
            return format_str % locals()
    return format_str % dict(symbol=symbols[0], value=n)


def bytes2file_response(bytes_info):
    yield bytes_info


def get_excel_template(header_list: List, selector_header_list: List, option_list: List[dict]):
    """
    工具方法：将需要导出的list数据转化为对应excel的二进制数据
    :param header_list: 表头数据列表
    :param selector_header_list: 需要设置为选择器格式的表头数据列表
    :param option_list: 选择器格式的表头预设的选项列表
    :return: 模板excel的二进制数据
    """
    # 创建Excel工作簿
    wb = Workbook()
    # 选择默认的活动工作表
    ws = wb.active

    # 设置表头文字
    headers = header_list

    # 设置表头背景样式为灰色，前景色为白色
    header_fill = PatternFill(start_color="ababab", end_color="ababab", fill_type="solid")

    # 将表头写入第一行
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        # 设置列宽度为16
        ws.column_dimensions[chr(64 + col_num)].width = 12
        # 设置水平居中对齐
        cell.alignment = Alignment(horizontal="center")

    # 设置选择器的预设选项
    options = option_list

    # 获取selector_header的字母索引
    for selector_header in selector_header_list:
        column_selector_header_index = headers.index(selector_header) + 1

        # 创建数据有效性规则
        header_option = []
        for option in options:
            if option.get(selector_header):
                header_option = option.get(selector_header)
        dv = DataValidation(type="list", formula1=f'"{",".join(header_option)}"')
        # 设置数据有效性规则的起始单元格和结束单元格
        dv.add(
            f"{get_column_letter(column_selector_header_index)}2:{get_column_letter(column_selector_header_index)}1048576"
        )
        # 添加数据有效性规则到工作表
        ws.add_data_validation(dv)

    # 保存Excel文件为字节类型的数据
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    return file.getvalue()


def get_filepath_from_url(url: str):
    """
    工具方法：根据请求参数获取文件路径
    :param url: 请求参数中的url参数
    :return: 文件路径
    """
    file_info = url.split("?")[1].split("&")
    task_id = file_info[0].split("=")[1]
    file_name = file_info[1].split("=")[1]
    task_path = file_info[2].split("=")[1]
    return os.path.join(setting.STATIC_DIR, task_path, task_id, file_name)


SchemaType = TypeVar("SchemaType", bound=BaseModel)


def get_form_dict(
    form: Optional[SchemaType] = None,
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    time_range: Optional[str] = None,
    **kwargs,
) -> Dict:
    form_dict = form.dict(exclude_none=True) if form else {}
    if (start is not None) and (end is not None) and (time_range is not None):
        form_dict[time_range] = (start, end)
    form_dict.update(kwargs)
    return form_dict


def create_dir(file_name: str) -> str:
    """创建文件夹"""

    base_path = get_current_directory()

    path = base_path + os.sep + file_name + os.sep  # 拼接日志文件夹的路径

    os.makedirs(path, exist_ok=True)  # 如果文件夹不存在就创建

    return path


def get_current_directory() -> AnyStr:
    """
    获取当前目录
    :return:
    """
    current_path = os.path.dirname(__file__)  # 获取当前文件夹
    # 获取当前文件夹的上一层文件
    return os.path.abspath(os.path.join(current_path, ".."))


def get_uuid() -> str:
    """生成uuid"""
    return str(uuid.uuid4())


def string_to_bool(string: str) -> bool:
    """字符串转bool变量"""
    return string.lower() != "False"


def get_array_by_str(ids_str: str):
    ids_list = ids_str.split(",")
    return [int(x) for x in ids_list]


def is_http(string: str) -> Optional[str]:
    """是否是http"""
    return string if is_url(string, allowed_schemes=["http", "https"]) else None


def is_url(string: str, allowed_schemes: Optional[List[str]] = None) -> bool:
    """判断是否是URL"""
    try:
        result = urlparse(string)
        if allowed_schemes:
            return all([result.scheme, result.netloc, result.scheme in allowed_schemes])
        else:
            return all([result.scheme, result.netloc])
    except ValueError:
        return False


def generate_id():
    import uuid

    uuid_ = uuid.uuid1()
    uuid_str = str(uuid_).replace("-", "")
    uuid_str = re.sub("[a-zA-Z]", "", uuid_str)
    last_12_chars = uuid_str[:12]
    return f"A{last_12_chars}"
